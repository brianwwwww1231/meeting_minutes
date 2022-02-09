from openpyxl import Workbook
wb = Workbook() # 先用Workbook() -> 建立物件，並且存進wb；type : Workbook

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells

company = input('請輸公司名稱： ')
products = input('請輸入本次申貸產品： ')
at = input('請輸入出席人員: ')
year = input('請輸入年： ')
month = input('請輸入月： ')
day = input('請輸入時間： ')
time = input('請輸入開會時間： ')
leader = input('請輸入主席名稱： ') 
record = input('請輸入紀錄者名稱： ')
w_t = '本公司擬向OOO商業銀行申請' # wording for topics
w_c = '本公司為配合實際業務需要，擬向OOO商業銀行申請' # wording for contents

ws.merge_cells('A1:K1')
ws.merge_cells('A2:K2')
ws.merge_cells('C10:K10') # 案由
ws.merge_cells('C11:K11') # 說明
ws.merge_cells('B7:K7') # 出席人員

# topic_1 董事會案由
tc_1 = w_t + products + '，期間均為壹年，授權董事長處理相關事宜。'
# meet_content_1 董事會議紀錄
mc_1 = w_c + products + '，期間均為壹年，均授權董事長全權處理，並代表本公司與該銀行簽立有關之契據及辦理該銀行要求之一切必要手續。此外，本公司若名下不動產設定抵押權予銀行之需要時，亦特別授權董事長全權處理，並代表本公司與該銀行簽立設定抵押權相關之契據及辦理該銀行要求之設定、變更、塗銷等必要手續。'

# topic_2 股東會案由
tc_2 = w_t + products + '，期間均為壹年，授權董事處理相關事宜。'
# meet_content_2 股東會議紀錄
mc_2 = w_c + products + '，期間均為壹年，均授權董事全權處理，並代表本公司與該銀行簽立有關之契據及辦理該銀行要求之一切必要手續。此外，本公司若名下不動產設定抵押權予銀行之需要時，亦特別授權董事全權處理，並代表本公司與該銀行簽立設定抵押權相關之契據及辦理該銀行要求之設定、變更、塗銷等必要手續。'

# topic_3 虧損案由
tc_3 = '本公司目前累積虧損已逹實收資本額二分之一以上，依據公司法第二一一條之規定，提請討論。'
# meet_content_3 虧損
mc_3 = w_c + products + '，惟目前累積虧損已逾實收資本額之二分之一，依據公司法第二一一條之規定提請討論，本公司近年來公司營運正常，股東權益已逐年穩定，未來公司將持續追求長期穩定獲利為主要目標，並逐年以獲利來改善公司整體財務結構，使公司能永續經營。'

def write(company, year, month, day, time, leader, record, at):
	ws['A1'] = company
	ws['A4'] = '一、時 間： ' 
	ws['B4'] ='中華民國'
	ws['C4'] = year
	ws['D4'] = '年'
	ws['E4'] = month
	ws['F4'] = '月'
	ws['G4'] = day
	ws['H4'] = '日'
	ws['J4'] = time
	ws['A5'] = '二、地 點： '
	ws['B5'] = '本公司會議室'
	ws['A6'] = '三、主 席： '
	ws['B6'] = leader
	ws['G6'] = '記 錄： '
	ws['H6'] = record
	ws['A7'] = '四、出席人員： '
	ws['B7'] = at
	ws['A8'] = '五、報告事項： '
	ws['B8'] = '(略)'
	ws['A9'] = '六、討論事項： '
	ws['B10'] = '案 由： '
	ws['B11'] = '說 明： '
	ws['B12'] = '決 議： '
	ws['C12'] = '經主席徵詢全體出席股東無異議照案通過。'
	ws['B13'] = '(餘略)'
	ws['A14'] = '七、散 會'

def save():
	wb.save("會議記錄.xlsx")
	print('記得先去資料夾換檔名，免得等下其他的動作蓋掉檔案！')

while True:
	meets = input('會議記錄型態([董事會]: 1 /[股東會]: 2/ [虧損>1/2]: 3)： ')
	if meets == '1':
		ws['A2'] = '董事會決議錄節錄'
		ws['C10'] = tc_1
		ws['C11'] = mc_1
		write(company, year, month, day, time, leader, record, at)
		save()
	elif meets == '2':
		check = input('是否重新輸入股東會議出席人員(y/n)： ')
		if check == 'y':
			at = input('請輸入出席人員: ')
			ws['A2'] = '股東會決議錄節錄'
			ws['C10'] = tc_2
			ws['C11'] = mc_2
			write(company, year, month, day, time, leader, record, at)
			save()
		elif check == 'n':
			ws['A2'] = '股東會決議錄節錄'
			ws['C10'] = tc_2
			ws['C11'] = mc_2
			write(company, year, month, day, time, leader, record, at)
			save()
		elif check == 'q':
			continue
		else:
			print('輸入錯誤，請重新輸入！')
	elif meets == '3':
		check = input('是否重新輸入股東會議出席人員(y/n)： ')
		if check == 'y':
			at = input('請輸入出席人員: ')
			ws['A2'] = '股東臨時會議事錄節錄'
			ws['C10'] = tc_3
			ws['C11'] = mc_3
			write(company, year, month, day, time, leader, record, at)
			save()
		elif check == 'n':
			ws['A2'] = '股東臨時會議事錄節錄'
			ws['C10'] = tc_3
			ws['C11'] = mc_3
			write(company, year, month, day, time, leader, record, at)
			save()
		elif check == 'q':
			continue
		else:
			print('輸入錯誤，請重新輸入！')
	elif meets == 'q':
		break
	else:
		print('輸入錯誤，請重新輸入！')



# # Rows can also be appended
# ws.append([1, 2, 3])

# Python types will automatically be converted
# import datetime
# ws['A20'] = datetime.datetime.now()

# Save the file
wb.save("sample_2.xlsx") # 再使用物件裡的功能「.xxx()」